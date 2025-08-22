import pandas as pd
import os
import numpy as np
import gc
import traceback

# Usa try/except para a importação do openpyxl
try:
    from openpyxl import load_workbook
except ImportError:
    # Implementação de fallback para contar linhas
    def count_rows_in_excel(filepath):
        """Fallback para contar linhas usando pandas"""
        try:
            # Verifica se o arquivo existe
            if not os.path.exists(filepath):
                return 0
            
            # Tenta ler apenas 1 linha para ver se o arquivo é válido
            test_df = pd.read_excel(filepath, nrows=1)
            
            # Se chegou aqui, o arquivo é válido, usa uma estimativa baseada no tamanho
            file_size = os.path.getsize(filepath)
            # Estimativa grosseira: ~100 bytes por linha
            estimated_rows = min(file_size // 100, 10000)
            return max(estimated_rows, 10)  # No mínimo 10 linhas
        except Exception as e:
            print(f"Erro ao estimar linhas: {e}")
            return 100  # Valor padrão seguro

def count_rows_in_excel(filepath):
    """Conta o número de linhas em um arquivo Excel sem carregar todo o conteúdo na memória."""
    try:
        # Verifica se load_workbook está disponível (definido acima)
        if 'load_workbook' in globals():
            # Carrega apenas as propriedades do workbook
            wb = load_workbook(filepath, read_only=True)
            sheet = wb.active
            # Retorna o número máximo de linhas
            row_count = sheet.max_row
            # Libera memória
            del wb
            gc.collect()
            return row_count
        else:
            # Usa o fallback definido acima
            return count_rows_in_excel(filepath)
    except Exception as e:
        print(f"Erro ao contar linhas: {e}")
        print(traceback.format_exc())
        
        # Fallback seguro: estima linhas com base no tamanho do arquivo
        try:
            file_size = os.path.getsize(filepath)
            # Estimativa grosseira: ~100 bytes por linha
            estimated_rows = min(file_size // 100, 10000)
            return max(estimated_rows, 10)  # No mínimo 10 linhas
        except:
            return 100  # Valor padrão muito seguro

def process_excel_in_chunks(filepath, clima_df, chunk_size=100, col_plantio="Data de Plantio", 
                           col_sfwd="05. SFWD", col_pfwd="06. PFWD"):
    """
    Processa um arquivo Excel em chunks para evitar esgotar a memória.
    Retorna um DataFrame com os resultados.
    """
    try:
        # Contar linhas para determinar o número de chunks necessários
        total_rows = count_rows_in_excel(filepath)
        if total_rows <= 1:  # Apenas o cabeçalho ou vazio
            print(f"Arquivo {filepath} vazio ou contém apenas cabeçalho.")
            return pd.DataFrame(), 0, 0, 0
    except Exception as e:
        print(f"Erro ao contar linhas, usando valor padrão: {e}")
        # Se falhar em contar linhas, use um valor padrão seguro
        total_rows = 1000
    
    print(f"Processando arquivo {filepath} com aproximadamente {total_rows} linhas em chunks de {chunk_size}")
    
    # Obter os nomes das colunas do arquivo (apenas a primeira linha)
    header_df = pd.read_excel(filepath, nrows=0)
    all_columns = header_df.columns.tolist()
    
    # Verificar se as colunas necessárias existem
    if col_plantio not in all_columns:
        print(f"Coluna '{col_plantio}' não encontrada no arquivo {filepath}")
        return pd.DataFrame(), 0, 0, 0
    
    if col_sfwd not in all_columns:
        print(f"Coluna '{col_sfwd}' não encontrada no arquivo {filepath}")
        return pd.DataFrame(), 0, 0, 0
    
    # Verificar se a coluna PFWD existe
    incluir_pfwd = col_pfwd in all_columns
    if incluir_pfwd:
        print(f"Coluna '{col_pfwd}' encontrada. Incluindo cálculos PFWD.")
    else:
        print(f"Coluna '{col_pfwd}' não encontrada. Ignorando cálculos PFWD.")
    
    # Lista para armazenar os resultados de cada chunk
    results = []
    erros = 0
    linhas_validas = 0
    gdu_alto = 0
    
    # Processar em chunks para economizar memória
    for chunk_start in range(0, total_rows, chunk_size):
        try:
            # Carregar um chunk do arquivo
            chunk = pd.read_excel(
                filepath, 
                skiprows=chunk_start if chunk_start == 0 else chunk_start+1,  # Pular cabeçalho nos chunks subsequentes
                nrows=chunk_size,
                engine='openpyxl',
                na_values=['NA', 'N/A', 'n/a', 'na', '', 'null', 'NULL', 'None', 'none'],
                keep_default_na=False
            )
            
            # Se é o primeiro chunk e não tem linhas após o cabeçalho, pular
            if chunk_start == 0 and chunk.shape[0] == 0:
                continue
                
            # Se não é o primeiro chunk, adicionar cabeçalhos corretos
            if chunk_start > 0:
                # Garantir que o chunk tenha os mesmos nomes de colunas que o cabeçalho original
                if len(chunk.columns) == len(all_columns):
                    chunk.columns = all_columns
            
            # Criar DataFrame de resultado para este chunk
            chunk_result = chunk.copy(deep=False)
            
            # Processar cada linha do chunk
            for i, row in chunk.iterrows():
                try:
                    # Converter as datas para o processamento interno
                    data_plantio = pd.to_datetime(row[col_plantio], dayfirst=True, errors='coerce')
                    data_sfwd = pd.to_datetime(row[col_sfwd], dayfirst=True, errors='coerce')
                    
                    if pd.isna(data_plantio) or pd.isna(data_sfwd):
                        chunk_result.at[i, 'dias'] = float('nan')
                        chunk_result.at[i, 'gdu_acumulado'] = float('nan')
                        erros += 1
                    else:
                        intervalo = clima_df[(clima_df['data'] > data_plantio) & (clima_df['data'] <= data_sfwd)].copy()
                        intervalo['GDU'] = ((intervalo['temp_min'] + intervalo['temp_max']) / 2) - 10
                        gdu_acumulado = intervalo['GDU'].sum()
                        dias = (data_sfwd - data_plantio).days
                        
                        chunk_result.at[i, 'dias'] = dias
                        chunk_result.at[i, 'gdu_acumulado'] = round(gdu_acumulado, 2)
                        
                        # Verificar se GDU está acima do limite
                        if gdu_acumulado > 1200:
                            gdu_alto += 1
                            
                        linhas_validas += 1
                    
                    # Processar PFWD se disponível
                    if incluir_pfwd:
                        data_pfwd = pd.to_datetime(row[col_pfwd], dayfirst=True, errors='coerce')
                        
                        if pd.isna(data_pfwd):
                            chunk_result.at[i, 'dias_pfwd'] = float('nan')
                            chunk_result.at[i, 'gdu_acumulado_pfwd'] = float('nan')
                            erros += 1
                        else:
                            intervalo_pfwd = clima_df[(clima_df['data'] > data_plantio) & (clima_df['data'] <= data_pfwd)].copy()
                            intervalo_pfwd['GDU'] = ((intervalo_pfwd['temp_min'] + intervalo_pfwd['temp_max']) / 2) - 10
                            gdu_pfwd = intervalo_pfwd['GDU'].sum()
                            dias_pfwd = (data_pfwd - data_plantio).days
                            
                            chunk_result.at[i, 'dias_pfwd'] = dias_pfwd
                            chunk_result.at[i, 'gdu_acumulado_pfwd'] = round(gdu_pfwd, 2)
                            
                            # Verificar se GDU está acima do limite
                            if gdu_pfwd > 1200:
                                gdu_alto += 1
                                
                            linhas_validas += 1
                
                except Exception as e:
                    print(f"Erro ao processar linha {i} no chunk {chunk_start}-{chunk_start+chunk_size}: {e}")
                    chunk_result.at[i, 'dias'] = float('nan')
                    chunk_result.at[i, 'gdu_acumulado'] = float('nan')
                    if incluir_pfwd:
                        chunk_result.at[i, 'dias_pfwd'] = float('nan')
                        chunk_result.at[i, 'gdu_acumulado_pfwd'] = float('nan')
                    erros += 1
            
            # Adicionar o resultado deste chunk à lista
            results.append(chunk_result)
            
            # Liberar memória após processar cada chunk
            del chunk
            gc.collect()
            
            print(f"Chunk {chunk_start}-{chunk_start+chunk_size} processado: {chunk_result.shape[0]} linhas")
            
        except Exception as e:
            print(f"Erro ao processar chunk {chunk_start}-{chunk_start+chunk_size}: {e}")
            continue
    
    # Se não há resultados, retornar DataFrame vazio
    if not results:
        return pd.DataFrame(), erros, linhas_validas, gdu_alto
    
    # Combinar todos os chunks em um único DataFrame
    try:
        final_result = pd.concat(results, ignore_index=True)
        print(f"Arquivo processado com sucesso: {final_result.shape[0]} linhas totais")
        
        # Limpar a lista de resultados para economizar memória
        del results
        gc.collect()
        
        return final_result, erros, linhas_validas, gdu_alto
    except Exception as e:
        print(f"Erro ao combinar chunks: {e}")
        return pd.DataFrame(), erros, linhas_validas, gdu_alto
